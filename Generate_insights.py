import argparse
import psycopg2
import pandas as pd
import json
import util
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

# Load configuration
with open("config.json", "r") as config_file:
    config = json.load(config_file)

# Configure logging
try:
    util.set_up_logging()
    util.log_info("###############")
    util.log_info("SQL TO EXCEL in Progress")
except Exception as e:
    util.log_info(f"Error occurred: {str(e)}")

# Database connection details
DB_PARAMS = {
    "dbname": config["dbname"],
    "user": config["user"],
    "password": config["password"],
    "host": config["host"],
    "port": config["port"],
}

# Argument parser setup
parser = argparse.ArgumentParser(
    description="Generate Excel report from SQL queries with dynamic month and year."
)
parser.add_argument(
    "--month", type=str, required=True, help="Month for the report (e.g., 'september')."
)
parser.add_argument(
    "--year", type=int, required=True, help="Year for the report (e.g., 2025)."
)
args = parser.parse_args()

# Extract month and year from arguments
month = args.month.lower()
year = args.year
month_year = f"{month}_{year}"  # Create the dynamic month_year string

# File paths with dynamic month and year
sql_file_path = "Sql_Analysis/monthly_insights.sql"
output_excel_file = f"REPORTS/{month.capitalize()}_{year}_REPORT.xlsx"


def read_sql_queries(file_path):
    """Read and split SQL queries from a file."""
    try:
        with open(file_path, "r") as file:
            sql_content = file.read()
        # Split queries based on ';' and remove extra whitespace
        sql_queries = [
            query.strip() for query in sql_content.split(";") if query.strip()
        ]
        return sql_queries
    except FileNotFoundError:
        util.log_info(f"SQL file not found at {file_path}. Please check the file path.")
        return []


def fetch_data_from_postgres(sql_query, connection):
    """Execute SQL query and return data as a DataFrame."""
    # Replace placeholder with actual month_year
    sql_query = sql_query.replace("{month_year}", month_year)

    try:
        df = pd.read_sql(sql_query, connection)
        return df
    except Exception as e:
        util.log_info(f"Error executing SQL query: {sql_query}\n{e}")
        return pd.DataFrame()


def format_excel_sheet(sheet):
    """Format the Excel sheet: set header color, row height, borders, and column width."""
    # Set the background color for the header row (row 1 in Excel, index 0 in Python)
    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    for cell in sheet[1]:  # Row 1 corresponds to the header
        cell.fill = header_fill

    # Set the row height for the header row
    sheet.row_dimensions[1].height = 20  # Row index 1 for the header

    # Create a border style
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Apply borders to all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )  # Center alignment and wrap text

    # Set column width to 250 pixels for all columns
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                pass
        sheet.column_dimensions[column_letter].width = (
            250 / 7.5
        )  # Convert pixels to Excel width units


def main():
    # Connect to the PostgreSQL database
    try:
        conn = psycopg2.connect(**DB_PARAMS)
        util.log_info("DB Connected for SQL to Excel conversion.")
    except Exception as e:
        util.log_info(f"Error connecting to database: {e}")
        return

    # Read SQL queries
    sql_queries = read_sql_queries(sql_file_path)
    if not sql_queries:
        util.log_info("No SQL queries to process. Exiting.")
        return

    sheet_names = [
        "Savings Rate",  # 1. Savings Rate
        "Recurring vs Non-Recurring",  # 2. Recurring vs. Non-Recurring Expenses
        "Expense Efficiency",  # 3. Expense Efficiency: Needs vs Wants
        "Highest Daily Expense",  # 4. Daily Expense Comparison: Highest Daily Expense
        "Lowest Daily Expense",  # 5. Daily Expense Comparison: Lowest Daily Expense
        "Weekly Expense Analysis",  # 6. Weekly Analysis
        "Unexpected Expenses",  # 7. Unexpected/Overbudget Expenses
        "Daily Spending Variance",  # 8. Daily Spending Variance from Budget
        "Opportunity Cost Analysis",  # 9. Opportunity Cost Analysis
        "Psychological Spending",  # 10. Psychological Spending Analysis
        "Needs Wants Investments",  # 11. Needs, Wants, and Investment Distribution
        "Pending Payments",  # 12. Payment Mode Status
        "1st 15 Days vs Last 15 Days",  # 13. 1st 15 Days vs Last 15 Days Expenses
        "Max Expense Category",  # 14. Maximum Expense Category and its Percentage
        "Top N Expense Categories",  # 15. Top N Expense Categories
        "Weekdays vs Weekends",  # 16. Weekdays vs Weekends Spending
        "Category-wise",  # 17. Category-wise Distribution by Day Type
    ]

    # Export each query result to a separate sheet in Excel
    with pd.ExcelWriter(output_excel_file, engine="openpyxl") as writer:
        for index, query in enumerate(sql_queries):
            df = fetch_data_from_postgres(query, conn)
            if not df.empty:
                # Use specific sheet names
                sheet_name = (
                    sheet_names[index]
                    if index < len(sheet_names)
                    else f"sheet{index + 1}"
                )
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                util.log_info(f"Query {index + 1} data written to {sheet_name}.")

                # Format the Excel sheet
                format_excel_sheet(writer.sheets[sheet_name])
            else:
                util.log_info(f"No data returned for query {index + 1}.")

    # Close the database connection
    conn.close()
    util.log_info("Database connection closed.")
    util.log_info(f"Data export completed. Check the file: {output_excel_file}")


if __name__ == "__main__":
    main()
